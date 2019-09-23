#!/usr/bin/env python3
"""
  yml2xls.py

  Uses information in a YAML source file to fill out a UNL Physics Dept.
  Requisition form. If the YAML file has more items than can fit on a
  single form, more forms are automatically added.

  Requires: openpyxl, PyYAML, Libreoffice (for pdf output)
"""
import io
from urllib.request import urlopen
import datetime
from os.path import splitext

import openpyxl
from openpyxl.writer.excel import save_virtual_workbook
import yaml


def chunks(d, n):
    lst = [(k, v) for (k, v) in d.items()]
    cs = []
    for i in range(0, len(lst), n):
        cs.append({k: v for (k, v) in lst[i : i + n]})
    return cs


class UNLRequisition:
    _blank_form_raw = None

    def __init__(self, src_filename):
        self.name = splitext(src_filename)[0]
        self.fnames = []

        with open(src_filename, "r") as f:
            self.src = yaml.safe_load(f)

    def fetch_empty_form(self):
        """
        Fetches an blank requisition form from the physics website
        """
        global _blank_form_raw
        blank_req_url = "http://www.unl.edu/physics/docs/Requisition2014.xlsx"
        if self._blank_form_raw is None:
            self._blank_form_raw = urlopen(blank_req_url).read()
        bio = io.BytesIO(self._blank_form_raw)
        self.workbook = openpyxl.load_workbook(bio)

    def populate_misc_fields(self):
        ws = self.workbook.active
        src = self.src

        ws["B11"] = src["vendor"].get("name", "")
        ws["B13"] = src["vendor"].get("address", "")
        city = src["vendor"].get("city", "")
        state = src["vendor"].get("state", "")
        zip = src["vendor"].get("zip", "")
        ws["B16"] = f"{city}, {state} {zip}"
        contact_name = src["vendor"].get("contact_name", "")
        contact_phone = src["vendor"].get("contact_phone", "")
        ws["B24"] = f"{contact_name}, {contact_phone}"
        ws["B22"] = src["vendor"].get("phone", "")
        ws["F22"] = src["vendor"].get("fax", "")
        ws["B26"] = src["vendor"].get("url", "")

        ws["E28"] = src.get("delivery_date", "")
        ws["D43"] = src.get("cost_object", "")
        ws["B45"] = src.get(
            "submission_date", datetime.date.today().strftime("%b. %d, %Y")
        )
        ws["C47"] = src["requestor_name"]
        ws["K47"] = src["requestor_phone"]
        ws["C49"] = src["supervisor_name"]

        ws["B18"] = src["use_for_project"]

    def populate_parts(self, parts):
        ws = self.workbook.active
        for i, (part, info) in enumerate(parts.items()):
            i = str(i + 32)
            ws["A" + i] = part
            ws["B" + i] = info.get("desc", "")
            ws["L" + i] = info.get("quantity", 1)
            ws["Q" + i] = info.get("unit_price", "N/A")

    def place_sheet_number(self, i, n):
        ws = self.workbook.active
        ws["A52"] = "Sheet {} of {}".format(i, n)

    def save_form(self):
        parts_chunked = chunks(self.src["items"], 10)
        for i, parts_chunk in enumerate(parts_chunked):
            self.fetch_empty_form()
            self.populate_parts(parts_chunk)
            self.populate_misc_fields()
            self.place_sheet_number(i + 1, len(parts_chunked))

            if len(parts_chunked) > 1:
                self.fnames.append(f"{self.name}_{i:02d}.xlsx")
            else:
                self.fnames.append(f"{self.name}.xlsx")
            with open(self.fnames[-1], "wb") as f:
                f.write(save_virtual_workbook(self.workbook))


def main():
    from argparse import ArgumentParser

    parser = ArgumentParser("yml2xls")
    parser.add_argument("input_file", help="A YAML file specifying the order items")
    parser.add_argument("--pdf", action="store_true")

    args = parser.parse_args()

    form = UNLRequisition(args.input_file)
    form.save_form()

    if args.pdf:
        from subprocess import call
        from os import remove

        print("Converting output to pdf")
        for fname in form.fnames:
            retcode = call(["libreoffice", "--convert-to", "pdf", fname])
            if retcode == 0:  # Success
                remove(fname)
            else:
                print(
                    (
                        "Failed to convert the form to pdf. "
                        "Is Libreoffice installed and available on PATH?"
                    )
                )


if __name__ == "__main__":
    main()
